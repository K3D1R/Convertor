from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import os, fnmatch

statuses = ['Зритель', 'Участник', 'Помощник - место проведения', 'Помощник организатора', '*Организатор', 'Экскурсовод', 'Корреспондент', 'Фотокорреспондент', 'Группа поддержки']
statuses_pl = ['зритель', 'участник', 'помощник - место проведения', 'помощник организатора', '*организатор', 'экскурсовод', 'корреспондент', 'фотокорреспондент', 'группа поддержки']
ekskurs = ['каб 17', 'депо', 'моделист', 'экскурсовод']
achievements = ['Зритель (всерос)', 'Участник (всерос)', '3 место (всерос)', '2 место (всерос)', '1 место (всерос)', 'Зритель (регион)', '3 место (регион)', '2 место (регион)', '1 место (регион)', 'Зритель (муницип)', '3 место (муницип)', '2 место (муницип)', '1 место (муницип)', 'Участник (регион)', 'Участник (муницип)', 'СЕРТИФИКАТ УЧАСТНИКА', '1 место','2 место','3 место']
achievements_low = ['зритель','участник','3 место','2 место','1 место']

levels = ['Всерос','МДЖД','Муницип','Регион','Междунар']
levels_low = ['всерос','муницип','регион','междунар']

print('Добро пожаловать в программу!')
turn = True
while turn:
    listOfFiles = os.listdir('.')
    pattern = "*.xlsx"
    for entry in listOfFiles:
        file = fnmatch.fnmatch(entry, pattern)
        if file:
            if entry[:2] == '~$':
                print(f"Если вы собираетесь работать с {entry.replace('~$', '')}, то необходимо закрыть его")
            print(entry)


    name = input('Введите имя файла: имяфайла.xlsx ')
    if name == 'вспв':
        turn = False
    else:
        wb = load_workbook(name,data_only=True)

        for sheet in wb:
            print(sheet.title)
        sheet_name = input()

        wsh = wb[sheet_name]
        wsh21 = wb.create_sheet(sheet_name + ' v2.1')
        inversion = bool(int(input('Поменять местами смены с табельными номерами? 0/1 ')))
        not_standart_template = bool(int(input('МЧ обычный или грамоты? 0/1 '))) #гыг

        #Создаём шапку
        title = str(wsh['A1'].value) #Здесь должно быть решение для случаев, когда число задано в названии, но его здесь нет
        wsh21['A1'] = title.replace('.', ' ')
        date = wsh['D1'].value
        if not date:
            print('Отсутствует дата мероприятия в ячейке D1!')
            date = input('Введите дату проведения мероприятия в формате дд.мм.гггг> ')
        wsh21['D1'] = date
        wsh21['E1'] = 'Уровень мероприятия:'
        level = wsh['G1'].value
        if level in levels:
            print(46)
            wsh21['G1'] = level
        elif level == 'МЖД':
            print(48)
            level = levels[3]
        elif not level:
            print('Не найден уровень мероприятия. Скопируйте нужынй уровень из приведённых ниже:')
            for st_level in levels:
                print(st_level)
            tr_level = input()
            if tr_level in levels:
                level = tr_level
            else:
                print('Неверно. Мероприятию будет присвоен уровень МДЖД')
                level = levels[1]

        else:
            print(63)
            moment = True
            i=1
            while moment:
                level=level[:i]
                i+=1
                if level in levels:
                    moment = False

                if i>=10:
                    moment=False
                    print('Не найден уровень мероприятия. Скопируйте нужынй уровень из приведённых ниже:')
                    for st_level in levels:
                        print(st_level)
                    tr_level = input()
                    if tr_level in levels:
                        level = tr_level
                    else:
                        print('Неверно. Мероприятию будет присвоен уровень МДЖД')
                        level = levels[1]
        wsh21['G1'] = level
        wsh21['A2'] = '№ п/п'
        wsh21['B2'] = 'ФИ южд'
        wsh21['C2'] = 'Смена'
        wsh21['D2'] = 'Таб. ном.'
        wsh21['E2'] = 'Примечание'

        if not not_standart_template:
            wsh21['F2'] = 'Кол-во км'
            wsh21['G2'] = 'Км за диплом'
            wsh21['H2'] = 'Бонус'
            wsh21['I2'] = 'Итог'
        else:
            wsh21['F2'] = 'Км за диплом'
            wsh21['I2'] = 'МДЖД/ВнеМДЖД'

        otvs = wsh['I1'].value
        if otvs == None:
            print('Необходимо заполнить поле I1: введите ответственного за заполнение ')
            wsh21['I1'] = input()
        else:
            wsh21['J1'] = otvs

        otvs_prep = wsh['J1'].value
        if not otvs_prep:
            wsh21['J1'] = 'Cмена №_ Фамилия И.О. преподавателя'
            print('Необходимо заполнить поле J1: введите ответственного преподавателя')
        else:
            wsh21['J1'] = otvs_prep
        #Конец шапки
        end_num = 0
        for col in wsh.iter_cols(max_col=1):
            for cell in col:
                end_num = cell.row
                if cell.value == None:
                    break
        end_num=end_num - 1
        #fontarea

        thin = Side(border_style="thin", color="000000")
        font_area_start = 'A1'
        font_area_end = f'M{end_num}'
        font_area = wsh21[font_area_start:font_area_end]
        ft = Font(name = 'Montserrat')
        for column in wsh21.columns:
            for cell in column:
                if cell.row==end_num:
                    continue
                else:
                    cell.font = ft
                    cell.border = Border(top=thin,left=thin,right=thin,bottom=thin)


        #Копирование фамилии имени
        for row in range(3, end_num):
            cell = f'B{row}'
            wsh21[cell] = wsh[cell].value

        if not inversion:
            #Копирование смены
            for row in range(3, end_num):
                cell = f'C{row}'
                wsh21[cell] = wsh[cell].value

            #Копирование табельных номеров
            for row in range(3, end_num):
                cell = f'D{row}'
                wsh21[cell]=wsh[cell].value
        else:
            # Копирование смены
            for row in range(3, end_num):
                cell = f'C{row}'
                cell_1 = f'D{row}'
                wsh21[cell] = wsh[cell_1].value

            # Копирование табельных номеров
            for row in range(3, end_num):
                cell = f'D{row}'
                cell_1 = f'C{row}'
                wsh21[cell] = wsh[cell_1].value

        #должности
        if not not_standart_template:
            print(163)
            for row in range(3, end_num):
                cell = f'E{row}'
                status = wsh[cell].value
                print(status)
                if status:
                    if status.lower() in statuses_pl:
                        wsh21[cell] = status
                    else:
                        if status in ekskurs:
                            wsh21[cell] = statuses[5]
                        elif status =='ведущий':
                            wsh21[cell]=statuses[3]
                        else:
                            status = status + ' '
                            if status.lower() in ekskurs:
                                wsh21[cell] = statuses[5]
                            else:
                                wsh21[cell]=statuses[1]
                else:
                    continue
        else:
            print(181)
            for row in range(3, end_num):
                cell = f'E{row}'
                status = wsh[cell].value
                if status in achievements:
                    wsh21[cell]=status
                elif status.lower() in achievements_low:
                    wsh21[cell]=f"{status.lower().title()} ({wsh21['G1'].value.lower()})"
                else:
                    if wsh21['G1'].value=='МДЖД':
                        wsh21[cell]=achievements[-4]
                    else:
                        wsh21[cell]=f"{achievements_low[1].title()} ({wsh21['G1'].value.lower()})"

        #километры
        for row in range(3, end_num):
            cell = f'F{row}'
            km = wsh[cell].value
            print(km)
            wsh21[cell].value=km

        if not not_standart_template:
            print(194)

            if wsh['I3'].value == None:
                for row in range(3, end_num):
                    cell = f'G{row}'
                    wsh21[cell] = '-'
            else:
                for row in range(3, end_num):
                    cell = f'G{row}'
                    wsh21[cell] = wsh[cell].value

                for row in range(3, end_num):
                    cell = f'I{row}'
                    wsh21[cell] = wsh[cell].value

            for row in range(3, end_num):
                cell = f'H{row}'
                cell_1 = f'G{row}'
                wsh21[cell]=wsh[cell_1].value

        else:
            print(213)

            if not wsh['I3'].value and not wsh['G3'].value:
                print(229)
                for row in range(3, end_num):
                    cell = f'F{row}'
                    wsh21[cell].value = wsh[cell].value

            elif not wsh['I3'].value:
                print(224)
                for row in range(3, end_num):
                    cell = f'F{row}'
                    cell_1 = f'G{row}'
                    wsh21[cell] = wsh[cell_1].valuу
            else:
                for row in range(3, end_num):
                    cell = f'F{row}'
                    wsh21[cell]=wsh[cell].value

            if wsh21['G1'].value == 'МДЖД':
                for row in range(3, end_num):
                    cell = f'I{row}'
                    wsh21[cell]=0
            else:
                for row in range(3, end_num):
                    cell = f'I{row}'
                    wsh21[cell]=1

        num = 1
        for row in range(3, end_num):
            cell = f'A{row}'
            wsh21[cell] = num
            num+=1


        #print('Пожалуйста, добавьте вручную таблицу с км по сменам после того, как программа завершит работу')
        wsh21['L1'] = 'Согласовано с куратором мероприятия'
        wsh21['L2'] = 'Итог по сменам'
        wsh21['L3'] = 'Смена'
        wsh21['L4'] = '1 смена'
        wsh21['L5'] = '2 смена'
        wsh21['L6'] = '3 смена'
        wsh21['L7'] = '4 смена'
        wsh21['M3'] = 'Кол-во км'

        if not not_standart_template:
            print(278)
            wsh21['M4'] = f'=SUMIF(C3:C{end_num}, 1, I3:I{end_num})'
            wsh21['M5'] = f'=SUMIF(C3:C{end_num}, 2, I3:I{end_num})'
            wsh21['M6'] = f'=SUMIF(C3:C{end_num}, 3, I3:I{end_num})'
            wsh21['M7'] = f'=SUMIF(C3:C{end_num}, 4, I3:I{end_num})'
        else:
            print(284)
            wsh21['M4'] = f'=SUMIF(C3:C{end_num}, 1, F3:F{end_num})'
            wsh21['M5'] = f'=SUMIF(C3:C{end_num}, 2, F3:F{end_num})'
            wsh21['M6'] = f'=SUMIF(C3:C{end_num}, 3, F3:F{end_num})'
            wsh21['M7'] = f'=SUMIF(C3:C{end_num}, 4, F3:F{end_num})'

        print(end_num)
        wb.save(name)